import openpyxl

wb = openpyxl.load_workbook(filename="./sf.xlsx")
sf = wb["TDSheet"]
gtd = wb["gtd"]
wb.active = sf


def main():
    for row in range(1, sf.max_row):
        if sf[f"B{row}"].value == 1:
            sf_begin = row + 1  # The first line of goods list in "TDSheet" sheet
            break

    for row in range(sf_begin, sf.max_row):
        if "Всего к оплате" in sf[f"B{row}"].value:
            sf_end = row  # The last line of goods list in "TDSheet" sheet
            break

    # Formatting name of goods in "sf" sheet
    for row in range(sf_begin, sf_end + 1):
        sf.cell(row=row, column=2).value = (
            str(" ".join(sf[f"B{row}"].value.split())).split(", Страна", 1)[0].strip()
        )

    # Formatting name of goods in "gtd" sheet
    for row in range(1, gtd.max_row + 1):
        gtd.cell(row=row, column=2).value = str(" ".join(gtd[f"B{row}"].value.split())).strip()
        if str(gtd[f"E{row}"].value).strip() != "None":
            gtd.cell(row=row, column=5).value = str(gtd[f"E{row}"].value).strip()
        else:
            gtd.cell(row=row, column=5).value = str("-")

    goods = {}  # Dict of all goods with "yes" - uniq and "no" not uniq goods in "gtd" sheet
    for row in range(1, gtd.max_row + 1):
        if gtd[f"B{row}"].value not in goods:
            goods[gtd[f"B{row}"].value] = {"gtd_num": gtd[f"E{row}"].value, "uniq_gtd": "yes"}
        elif (
            gtd[f"B{row}"].value in goods
            and goods[gtd[f"B{row}"].value]["gtd_num"] != gtd[f"E{row}"].value
        ):
            goods[gtd[f"B{row}"].value].update({"uniq_gtd": "no"})

    notuniq_goods_tmp = goods.copy()  # Temporary dict of not uniq goods in "gtd" sheet
    uniq_goods = goods.copy()  # Dict of uniq goods in "gtd" sheet
    for val in goods:
        if goods[val]["uniq_gtd"] == "yes":
            del notuniq_goods_tmp[val]
        else:
            del uniq_goods[val]

    notuniq_goods = {}  # Dict of not uniq goods in "gtd" sheet with gtds and quantities
    for row in range(1, gtd.max_row + 1):
        if gtd[f"B{row}"].value in notuniq_goods_tmp:
            if gtd[f"B{row}"].value not in notuniq_goods:
                notuniq_goods[gtd[f"B{row}"].value] = {gtd[f"E{row}"].value: gtd[f"D{row}"].value}
            else:
                if gtd[f"E{row}"].value not in notuniq_goods[gtd[f"B{row}"].value]:
                    notuniq_goods[gtd[f"B{row}"].value].update(
                        {gtd[f"E{row}"].value: gtd[f"D{row}"].value}
                    )
                else:
                    new_val = (
                        notuniq_goods[gtd[f"B{row}"].value][gtd[f"E{row}"].value]
                        + gtd[f"D{row}"].value
                    )
                    notuniq_goods[gtd[f"B{row}"].value].update({gtd[f"E{row}"].value: new_val})

    # Inserting gtds from "gtd" sheet into "TDSheet" sheet
    for row in range(sf_begin, sf_end):
        if sf[f"B{row}"].value in uniq_goods:
            sf.cell(row=row, column=32).value = uniq_goods[sf[f"B{row}"].value]["gtd_num"]
        elif sf[f"B{row}"].value not in uniq_goods and sf[f"B{row}"].value not in notuniq_goods:
            sf.cell(row=row, column=32).value = "!!!!! ТОВАР НЕ НАЙДЕН !!!!!"
        else:
            for key, val in notuniq_goods[sf[f"B{row}"].value].items():
                if (
                    notuniq_goods[sf[f"B{row}"].value][key] == sf[f"J{row}"].value
                    and not sf[f"AF{row}"].value
                ):
                    sf.cell(row=row, column=32).value = key
                    notuniq_goods[sf[f"B{row}"].value].update({key: 0})

    wb.remove(gtd)
    wb.save("./sf_gtd.xlsx")


if __name__ == "__main__":
    main()
